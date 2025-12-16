#!/usr/bin/env python3
"""
Absence Data Aggregation Tool
Aggregates 12 monthly absence Excel files into yearly totals per employee.

Usage:
    python3 aggregate_absence.py

Input: 01.2568.xlsx through 12.2568.xlsx (in current directory)
Output: absence-summary-2568.xlsx
"""

import pandas as pd
import glob
import re
from difflib import SequenceMatcher


def extract_name_key_and_notes(full_name):
    """
    Extract matching key and notes from full name string.
    Key = prefix|firstname|lastname|nickname (ignores extra parts like middle names)

    Example: "นาง CHHUN ORNG LY (รี)/ลาออก 27/03"
    -> key: "นาง|CHHUN|ORNG|รี", display_name: "นาง CHHUN ORNG (รี)", note: "ลาออก 27/03"
    """
    if not full_name or pd.isna(full_name):
        return None, None, None

    full_name = str(full_name).strip()

    # Extract notes after /
    note = None
    match = re.search(r'/(?=[ก-๙a-zA-Z])', full_name)
    if match:
        note = full_name[match.start()+1:].strip()
        name_part = full_name[:match.start()].strip()
    else:
        name_part = full_name

    # Extract nickname (Thai in parentheses)
    nick_match = re.search(r'\(([ก-๙]+)\)', name_part)
    nickname = nick_match.group(1) if nick_match else ''

    # Remove nickname from name
    name_clean = re.sub(r'\s*\([ก-๙]+\)\s*', '', name_part).strip()
    parts = name_clean.split()

    if not parts:
        return None, None, None

    first_part = parts[0]

    # Detect if Thai or Foreign name pattern
    if first_part in ['นาย', 'นาง', 'นางสาว']:
        # Foreign name: นาง FIRSTNAME LASTNAME [EXTRA...]
        prefix = first_part
        firstname = parts[1] if len(parts) > 1 else ''
        lastname = parts[2] if len(parts) > 2 else ''
    else:
        # Thai name: นายFirstname Lastname
        if first_part.startswith('นางสาว'):
            prefix, firstname = 'นางสาว', first_part[6:]
        elif first_part.startswith('นาง'):
            prefix, firstname = 'นาง', first_part[3:]
        elif first_part.startswith('นาย'):
            prefix, firstname = 'นาย', first_part[3:]
        else:
            prefix, firstname = '', first_part
        lastname = parts[1] if len(parts) > 1 else ''

    # Create matching key (ignore nickname - same person might have different nicknames)
    key = f'{prefix}|{firstname}|{lastname}'

    # Create display name (always add space after prefix, include nickname if available)
    if nickname:
        display_name = f'{prefix} {firstname} {lastname} ({nickname})'
    else:
        display_name = f'{prefix} {firstname} {lastname}'

    # Clean up extra spaces
    display_name = ' '.join(display_name.split())

    return key, display_name, note


def find_monthly_files():
    """Find all XX.2568.xlsx files in current directory"""
    files = glob.glob('[0-1][0-9].2568.xlsx')
    return sorted(files)


def parse_value(val):
    """Handle missing/empty values"""
    if pd.isna(val) or val == '' or str(val).strip() in ['-', ' - ']:
        return 0
    try:
        return float(val)
    except:
        return 0


def similarity_ratio(s1, s2):
    """Calculate string similarity (0-1)"""
    return SequenceMatcher(None, s1.lower(), s2.lower()).ratio()


def normalize_name_parts(name_str):
    """Extract and normalize name parts (prefix, firstname, lastname)"""
    if not name_str or pd.isna(name_str):
        return None, None, None

    name_str = str(name_str).split('/')[0].strip()  # Remove notes
    name_str = re.sub(r'\s*\([ก-๙]+\)\s*', '', name_str).strip()  # Remove nickname

    parts = name_str.split()
    if not parts:
        return None, None, None

    first_part = parts[0]
    if first_part in ['นาย', 'นาง', 'นางสาว']:
        prefix = first_part
        firstname = parts[1] if len(parts) > 1 else ''
        lastname = parts[2] if len(parts) > 2 else ''
    else:
        # Thai name with prefix merged
        if first_part.startswith('นางสาว'):
            prefix, firstname = 'นางสาว', first_part[6:]
        elif first_part.startswith('นาง'):
            prefix, firstname = 'นาง', first_part[3:]
        elif first_part.startswith('นาย'):
            prefix, firstname = 'นาย', first_part[3:]
        else:
            prefix, firstname = '', first_part
        lastname = parts[1] if len(parts) > 1 else ''

    return prefix.strip(), firstname.strip(), lastname.strip()


def find_fuzzy_match(prefix, firstname, lastname, employee_map, threshold=0.85):
    """
    Find fuzzy match in employee_map for potential typos/variations.
    Returns matching key if found, None otherwise.
    """
    for key, emp in employee_map.items():
        parts = key.split('|')
        if len(parts) >= 3:
            existing_prefix, existing_firstname, existing_lastname = parts[0], parts[1], parts[2]

            # Prefix must match exactly
            if prefix != existing_prefix:
                continue

            # Firstname must match exactly or very closely
            fname_sim = similarity_ratio(firstname, existing_firstname)
            if fname_sim < 0.95:  # Firstname must be nearly identical
                continue

            # Lastname can have typos - check similarity
            lname_sim = similarity_ratio(lastname, existing_lastname)
            if lname_sim >= threshold:
                return key

    return None


def load_monthly_file(filepath):
    """Load one monthly file, skip header rows"""
    df = pd.read_excel(filepath, skiprows=3)
    return df


def extract_absence_data(df):
    """Extract employee info + absence data from one month"""
    employees = []

    for _, row in df.iterrows():
        emp_id = row.iloc[0]      # รหัส (employee ID)
        full_name = row.iloc[1]   # ชื่อ-สกุล (may include notes after /)
        position = row.iloc[2]    # ตำแหน่ง
        department = row.iloc[3]  # แผนก
        pay_type = row.iloc[4]    # ประเภท

        # Skip if no name
        if pd.isna(full_name) or str(full_name).strip() == '':
            continue

        # Extract key (for matching), display name, and notes
        key, display_name, note = extract_name_key_and_notes(full_name)
        if not key:
            continue

        # Clean employee ID
        emp_id_str = str(emp_id).strip() if pd.notna(emp_id) else ''

        # Extract nickname for ID+nickname matching
        full_name_clean = str(full_name).split('/')[0].strip()
        nick_match = re.search(r'\(([ก-๙]+)\)', full_name_clean)
        nickname = nick_match.group(1) if nick_match else ''

        # Create matching key: use name-based key for fuzzy matching
        primary_key = key

        # Combine first half (cols 5-21) + second half (cols 22-38)
        monthly_totals = []
        for i in range(17):
            first_half = parse_value(row.iloc[5 + i])
            second_half = parse_value(row.iloc[22 + i])
            monthly_totals.append(first_half + second_half)

        employees.append({
            'primary_key': primary_key,
            'name_key': key,
            'emp_id': emp_id_str,
            'nickname': nickname,
            'display_name': display_name,
            'note': note,
            'position': position,
            'department': department,
            'payType': pay_type,
            'totals': monthly_totals
        })

    return employees


def aggregate_yearly_totals(all_months_data):
    """Aggregate absence data with fuzzy name matching for typos"""
    employee_map = {}

    for month_data in all_months_data:
        for emp in month_data:
            key = emp['primary_key']
            display_name = emp['display_name']

            # Try fuzzy match if exact key not found
            matched_key = key
            merge_reason = None
            if key not in employee_map:
                # Try fuzzy matching
                parts = key.split('|')
                if len(parts) >= 3:
                    prefix, firstname, lastname = parts[0], parts[1], parts[2]
                    fuzzy_key = find_fuzzy_match(prefix, firstname, lastname, employee_map, threshold=0.80)
                    if fuzzy_key:
                        matched_key = fuzzy_key
                        # Record why we merged (fuzzy match)
                        merge_reason = f"Fuzzy: {key} → {fuzzy_key}"

            # Initialize if first time seeing this key
            if matched_key not in employee_map:
                employee_map[matched_key] = {
                    'name': display_name,
                    'emp_ids': set(),
                    'notes': set(),
                    'original_names': set(),
                    'merge_reasons': set(),
                    'position': emp['position'],
                    'department': emp['department'],
                    'payType': emp['payType'],
                    'totals': [0] * 17
                }

            # Track original name
            employee_map[matched_key]['original_names'].add(display_name)

            # Track merge reason if fuzzy matched
            if merge_reason:
                employee_map[matched_key]['merge_reasons'].add(merge_reason)

            # Add employee ID if exists
            if emp['emp_id']:
                employee_map[matched_key]['emp_ids'].add(emp['emp_id'])

            # Add note if exists
            if emp['note']:
                employee_map[matched_key]['notes'].add(emp['note'])

            # Sum this month's totals
            for i in range(17):
                employee_map[matched_key]['totals'][i] += emp['totals'][i]

    # Convert sets to strings
    for emp in employee_map.values():
        emp['notes'] = ' | '.join(sorted(emp['notes'])) if emp['notes'] else ''
        emp['original_names'] = ' | '.join(sorted(emp['original_names'])) if emp['original_names'] else ''
        emp['merge_reasons'] = ' | '.join(sorted(emp['merge_reasons'])) if emp['merge_reasons'] else ''

        # Employee ID handling
        ids = sorted(emp['emp_ids'])
        if len(ids) == 0:
            emp['emp_id'] = ''
        elif len(ids) == 1:
            emp['emp_id'] = ids[0]
        else:
            emp['emp_id'] = ' | '.join(ids)

        del emp['emp_ids']

    # Sort by name
    return sorted(employee_map.values(), key=lambda x: x['name'])


def create_output_dataframe(aggregated_data):
    """Convert aggregated data to pandas DataFrame for export"""

    # Column headers (Thai + English)
    columns = [
        'รหัส (EmpID)',
        'ชื่อ-สกุล (Name)',
        'หมายเหตุ (Notes)',
        'ตำแหน่ง (Position)',
        'แผนก (Department)',
        'ประเภท (PayType)',
        'วันทำงาน (Work Days)',
        'ขาดงาน (Absent)',
        'ลากิจ (Personal Leave)',
        'ป่วยมีใบรพ. (Sick w/Cert)',
        'ป่วยไม่มีรพ. (Sick w/o Cert)',
        'ลาคลอด (Maternity)',
        'ลืมสแกนนิ้ว/มาสาย (Late Grace)',
        'มาสายเกิน (Late Penalty)',
        'ลาOT (OT Leave)',
        'ให้หยุด/พักงาน (Suspension)',
        'พักร้อน (Annual Leave)',
        'OT 2.5 ชม',
        'OT >2.5 ชม',
        'ทำงานวันหยุด (Holiday Work)',
        'OT วันหยุด (Holiday OT)',
        'กะดึก (Night Shift)',
        'ควบคุม 2 เครื่อง (Multi-Machine)'
    ]

    # Build rows
    rows = []
    for emp in aggregated_data:
        row = [
            emp['emp_id'],
            emp['name'],
            emp['notes'],
            emp['position'],
            emp['department'],
            emp['payType']
        ] + emp['totals']
        rows.append(row)

    df = pd.DataFrame(rows, columns=columns)
    return df


def calculate_summary_stats(aggregated_data, all_months_data):
    """Calculate summary statistics with raw file comparison for validation"""
    # Absence type columns
    absence_types = [
        'วันทำงาน (Work Days)',
        'ขาดงาน (Absent)',
        'ลากิจ (Personal Leave)',
        'ป่วยมีใบรพ. (Sick w/Cert)',
        'ป่วยไม่มีรพ. (Sick w/o Cert)',
        'ลาคลอด (Maternity)',
        'ลืมสแกนนิ้ว/มาสาย (Late Grace)',
        'มาสายเกิน (Late Penalty)',
        'ลาOT (OT Leave)',
        'ให้หยุด/พักงาน (Suspension)',
        'พักร้อน (Annual Leave)',
        'OT 2.5 ชม',
        'OT >2.5 ชม',
        'ทำงานวันหยุด (Holiday Work)',
        'OT วันหยุด (Holiday OT)',
        'กะดึก (Night Shift)',
        'ควบคุม 2 เครื่อง (Multi-Machine)'
    ]

    # Calculate raw totals from all monthly files (before merging)
    raw_totals = [0] * 17
    for month_data in all_months_data:
        for emp in month_data:
            for i in range(17):
                raw_totals[i] += emp['totals'][i]

    # Calculate aggregated totals
    aggregated_totals = [0] * 17
    for emp in aggregated_data:
        for i in range(17):
            aggregated_totals[i] += emp['totals'][i]

    # Add header info
    summary = []

    # Count total raw records before merging
    total_raw_records = sum(len(month) for month in all_months_data)
    total_merged_employees = len(aggregated_data)
    records_merged = total_raw_records - total_merged_employees

    summary.append({
        'Absence Type': '[VALIDATION INFO]',
        'Raw Files Total': '',
        'Summary Total': '',
        'Match?': '',
        'Notes': ''
    })
    summary.append({
        'Absence Type': f'Total monthly records processed',
        'Raw Files Total': total_raw_records,
        'Summary Total': total_merged_employees,
        'Match?': '✓ OK',
        'Notes': f'{records_merged} merged'
    })
    summary.append({
        'Absence Type': '[YEARLY TOTALS BY ABSENCE TYPE - COMPARE COLUMNS]',
        'Raw Files Total': '',
        'Summary Total': '',
        'Match?': '',
        'Notes': 'Raw Files Total = Raw data before merging | Summary Total = Your final numbers'
    })
    summary.append({
        'Absence Type': '',
        'Raw Files Total': '',
        'Summary Total': '',
        'Match?': '',
        'Notes': ''
    })

    # Add absence type totals with comparison
    for i, absence_type in enumerate(absence_types):
        raw_total = raw_totals[i]
        summary_total = aggregated_totals[i]
        match = '✓' if raw_total == summary_total else '❌ DIFF!'
        summary.append({
            'Absence Type': absence_type,
            'Raw Files Total': raw_total,
            'Summary Total': summary_total,
            'Match?': match,
            'Notes': '' if raw_total == summary_total else f'Difference: {raw_total - summary_total}'
        })

    return summary




def create_suspicious_sheet(df):
    """Create a sheet highlighting suspicious/problematic records with separate flag columns"""
    suspicious_records = []

    for idx, row in df.iterrows():
        emp_id = row['รหัส (EmpID)']
        name = row['ชื่อ-สกุล (Name)']
        notes = row['หมายเหตุ (Notes)']

        # Flag 1: Multiple IDs (job change)
        flag_multiple_ids = '⚠ YES' if (emp_id and '|' in str(emp_id)) else ''

        # Flag 2: Name has "/" in it (incomplete merging)
        flag_merged_name = '⚠ YES' if (name and '/' in str(name)) else ''

        # Flag 3: Has notes about ลาออก (quit)
        flag_quit = '⚠ YES' if (notes and 'ลาออก' in str(notes)) else ''

        # Flag 4: Has notes about เริ่มใหม่ (restarted)
        flag_restart = '⚠ YES' if (notes and 'เริ่มใหม่' in str(notes)) else ''

        # Flag 5: Has notes about ย้ายมา (transferred in)
        flag_transfer = '⚠ YES' if (notes and 'ย้ายมา' in str(notes)) else ''

        # Add to suspicious list if any flags
        if any([flag_multiple_ids, flag_merged_name, flag_quit, flag_restart, flag_transfer]):
            suspicious_records.append({
                'รหัส (ID)': emp_id,
                'ชื่อ-สกุล (Name)': name,
                'Multiple IDs?': flag_multiple_ids,
                'Merged Name?': flag_merged_name,
                'Quit (ลาออก)?': flag_quit,
                'Restart (เริ่มใหม่)?': flag_restart,
                'Transfer (ย้ายมา)?': flag_transfer,
                'หมายเหตุ (Notes)': notes
            })

    return pd.DataFrame(suspicious_records)


def create_executive_summary(aggregated_data, suspicious_df, all_months_data):
    """
    Create CEO-level executive summary sheet (30-second overview)

    Args:
        aggregated_data: List of employee dicts with yearly totals
        suspicious_df: DataFrame from create_suspicious_sheet()
        all_months_data: List of monthly employee data

    Returns:
        DataFrame with executive summary table
    """
    absence_types = [
        'วันทำงาน (Work Days)',
        'ขาดงาน (Absent)',
        'ลากิจ (Personal Leave)',
        'ป่วยมีใบรพ. (Sick w/Cert)',
        'ป่วยไม่มีรพ. (Sick w/o Cert)',
        'ลาคลอด (Maternity)',
        'ลืมสแกนนิ้ว/มาสาย (Late Grace)',
        'มาสายเกิน (Late Penalty)',
        'ลาOT (OT Leave)',
        'ให้หยุด/พักงาน (Suspension)',
        'พักร้อน (Annual Leave)',
        'OT 2.5 ชม',
        'OT >2.5 ชม',
        'ทำงานวันหยุด (Holiday Work)',
        'OT วันหยุด (Holiday OT)',
        'กะดึก (Night Shift)',
        'ควบคุม 2 เครื่อง (Multi-Machine)'
    ]

    summary = []

    # [SECTION 1: PERIOD & SCOPE]
    total_raw_records = sum(len(month) for month in all_months_data)
    total_employees = len(aggregated_data)
    records_merged = total_raw_records - total_employees

    summary.append({'Metric': '[PERIOD & SCOPE]', 'Value': ''})
    summary.append({'Metric': 'Data Period', 'Value': 'January - July 2568'})
    summary.append({'Metric': 'Total Unique Employees', 'Value': total_employees})
    summary.append({'Metric': 'Records Merged (deduplicates)', 'Value': records_merged})
    summary.append({'Metric': '', 'Value': ''})

    # [SECTION 2: WORKFORCE OVERVIEW]
    summary.append({'Metric': '[WORKFORCE OVERVIEW]', 'Value': ''})

    # Calculate totals
    work_days_total = sum(emp['totals'][0] for emp in aggregated_data)
    summary.append({'Metric': 'Total Work Days', 'Value': int(work_days_total)})

    # Suspicious counts
    total_suspicious = len(suspicious_df)
    multiple_ids = (suspicious_df['Multiple IDs?'] == '⚠ YES').sum() if len(suspicious_df) > 0 else 0
    quits = (suspicious_df['Quit (ลาออก)?'] == '⚠ YES').sum() if len(suspicious_df) > 0 else 0
    transfers = (suspicious_df['Transfer (ย้ายมา)?'] == '⚠ YES').sum() if len(suspicious_df) > 0 else 0

    summary.append({'Metric': 'Employees Requiring Review', 'Value': total_suspicious})
    summary.append({'Metric': '  └─ Job Changes (Multiple IDs)', 'Value': int(multiple_ids)})
    summary.append({'Metric': '  └─ Employee Quits', 'Value': int(quits)})
    summary.append({'Metric': '  └─ System Transfers', 'Value': int(transfers)})
    summary.append({'Metric': '', 'Value': ''})

    # [SECTION 3: TOP ABSENCE CATEGORIES]
    summary.append({'Metric': '[TOP ABSENCE CATEGORIES]', 'Value': ''})

    # Calculate absence totals
    absence_totals = [0] * 17
    for emp in aggregated_data:
        for i in range(17):
            absence_totals[i] += emp['totals'][i]

    # Create list of (type_name, total, index) and sort by total descending
    absence_list = [(absence_types[i], absence_totals[i], i) for i in range(17)]
    absence_list = [x for x in absence_list if x[1] > 0]  # Filter out zeros
    absence_list.sort(key=lambda x: x[1], reverse=True)

    # Show top 7 absence types with percentage
    for absence_name, total, idx in absence_list[:7]:
        pct = (total / work_days_total * 100) if work_days_total > 0 else 0
        summary.append({
            'Metric': absence_name,
            'Value': f'{total:.1f} days ({pct:.2f}%)'
        })

    summary.append({'Metric': '', 'Value': ''})

    # [SECTION 4: DEPARTMENT CONCENTRATION]
    summary.append({'Metric': '[DEPARTMENT CONCENTRATION (TOP 5)]', 'Value': ''})

    # Group by department
    dept_counts = {}
    for emp in aggregated_data:
        dept = emp.get('department', 'Unknown')
        dept_counts[dept] = dept_counts.get(dept, 0) + 1

    # Sort by count descending and take top 5
    top_depts = sorted(dept_counts.items(), key=lambda x: x[1], reverse=True)[:5]
    other_count = sum(count for _, count in sorted(dept_counts.items(), key=lambda x: x[1], reverse=True)[5:])

    for dept_name, count in top_depts:
        pct = (count / total_employees * 100)
        summary.append({'Metric': dept_name, 'Value': f'{count} employees ({pct:.1f}%)'})

    if other_count > 0:
        other_depts_count = len(dept_counts) - 5
        pct = (other_count / total_employees * 100)
        summary.append({
            'Metric': f'Other ({other_depts_count} departments)',
            'Value': f'{other_count} employees ({pct:.1f}%)'
        })

    summary.append({'Metric': '', 'Value': ''})

    # [SECTION 5: KEY INSIGHTS]
    summary.append({'Metric': '[KEY INSIGHTS]', 'Value': ''})

    # Calculate insights
    unexcused_absent = absence_totals[1]  # ขาดงาน (Absent)
    unexcused_pct = (unexcused_absent / work_days_total * 100) if work_days_total > 0 else 0

    personal_leave = absence_totals[2]  # ลากิจ (Personal Leave)
    personal_pct = (personal_leave / work_days_total * 100) if work_days_total > 0 else 0

    if unexcused_pct < 0.1:
        summary.append({
            'Metric': '✓ Low Compliance Risk',
            'Value': f'Unexcused absence only {unexcused_pct:.3f}% - excellent'
        })
    else:
        summary.append({
            'Metric': '⚠ Compliance Alert',
            'Value': f'Unexcused absence {unexcused_pct:.2f}% - review needed'
        })

    if total_suspicious > 0:
        suspicious_pct = (total_suspicious / total_employees * 100)
        summary.append({
            'Metric': '⚠ HR Review Required',
            'Value': f'{total_suspicious} employees ({suspicious_pct:.1f}%) - see Suspicious sheet'
        })

    if personal_pct > 2:
        summary.append({
            'Metric': 'ℹ High Personal Leave',
            'Value': f'{personal_pct:.2f}% of work days - planned absences dominate'
        })

    summary.append({'Metric': '✓ Data Quality', 'Value': f'{records_merged} duplicates merged, verified'})

    return pd.DataFrame(summary)


def create_merged_names_sheet(df, aggregated_data, all_months_data):
    """Create sheet showing ALL merged employees - by ID or by name algorithm"""
    merged_list = []

    # Month labels
    month_labels = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']

    for emp in aggregated_data:
        emp_id = emp.get('emp_id', '')
        original_names = emp.get('original_names', '')
        merge_reasons = emp.get('merge_reasons', '')

        # Check if merged: multiple IDs OR multiple original names OR has merge reasons
        has_multiple_ids = emp_id and '|' in str(emp_id)
        has_multiple_names = original_names and '|' in str(original_names)
        has_fuzzy_merge = bool(merge_reasons)

        if has_multiple_ids or has_multiple_names or has_fuzzy_merge:
            # Get ID for each month
            ids_list = [id_str.strip() for id_str in str(emp_id).split('|')] if emp_id else []

            month_ids = {}
            for m_idx, month_data in enumerate(all_months_data):
                month_label = month_labels[m_idx] if m_idx < len(month_labels) else f'M{m_idx+1}'
                found_ids = []
                for monthly_emp in month_data:
                    m_id = monthly_emp.get('emp_id', '').strip()
                    if m_id in ids_list:
                        found_ids.append(m_id)
                month_ids[month_label] = ' | '.join(sorted(found_ids)) if found_ids else '-'

            # Determine merge type
            if has_fuzzy_merge:
                merge_type = 'Fuzzy 80%'
            elif has_multiple_ids:
                merge_type = 'ID Change'
            else:
                merge_type = 'Name Variation'

            row = {
                'Final Name': emp['name'],
                'Original Names': original_names,
                'Merge Type': merge_type,
            }

            # Add each month's ID as separate column
            for m_idx in range(len(all_months_data)):
                month_label = month_labels[m_idx] if m_idx < len(month_labels) else f'M{m_idx+1}'
                row[month_label] = month_ids.get(month_label, '-')

            merged_list.append(row)

    # If no merged records, create empty structure
    if not merged_list:
        row = {
            'Final Name': 'No merged employees',
            'Original Names': '',
            'Merge Type': '',
        }
        for m_idx in range(len(all_months_data)):
            month_label = month_labels[m_idx] if m_idx < len(month_labels) else f'M{m_idx+1}'
            row[month_label] = ''
        merged_list.append(row)

    return pd.DataFrame(merged_list)


def export_to_excel(df, summary_df, aggregated_data, all_months_data, filename='absence-summary-2568.xlsx'):
    """Export DataFrame to Excel with executive summary, suspicious, summary and employees sheets"""
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment, PatternFill, Font
    from openpyxl.worksheet.datavalidation import DataValidation

    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        # Create suspicious sheet first (needed for executive summary calculation)
        suspicious_df = create_suspicious_sheet(df)

        # Write executive summary sheet (FIRST sheet for CEO view)
        executive_df = create_executive_summary(aggregated_data, suspicious_df, all_months_data)
        executive_df.to_excel(writer, sheet_name='Executive Summary', index=False)

        # Write suspicious records sheet
        suspicious_df.to_excel(writer, sheet_name='Suspicious', index=False)

        # Write merged names sheet (showing who was merged and all original names)
        merged_names_df = create_merged_names_sheet(df, aggregated_data, all_months_data)
        merged_names_df.to_excel(writer, sheet_name='Merged Names', index=False)

        # Write summary sheet
        summary_df.to_excel(writer, sheet_name='Summary', index=False)

        # Write detailed data
        df.to_excel(writer, sheet_name='Employees', index=False)

        # Auto-adjust column widths and style for all sheets
        for sheet_name in writer.sheets:
            ws = writer.sheets[sheet_name]
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 for readability
                ws.column_dimensions[column_letter].width = adjusted_width

                # Different alignment per sheet
                for cell in column:
                    if sheet_name == 'Executive Summary':
                        # Left-align for readability in summary sheet
                        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    else:
                        # Center align for detailed sheets
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            # Style Executive Summary sheet - bold headers, larger font
            if sheet_name == 'Executive Summary':
                section_header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
                section_header_font = Font(bold=True, size=12, color='000000')
                regular_font = Font(size=11)

                for row_idx, row in enumerate(ws.iter_rows(min_row=1), 1):
                    for cell in row:
                        cell.font = regular_font
                        if cell.value and '[' in str(cell.value) and ']' in str(cell.value):
                            cell.font = section_header_font
                            cell.fill = section_header_fill

            # Style Suspicious sheet - bold header, bold red text for flags
            elif sheet_name == 'Suspicious':
                bold_font = Font(bold=True, size=12)
                red_bold_font = Font(color='FF0000', bold=True, size=12)

                # Bold header row with larger font
                for cell in ws[1]:
                    if cell.value:
                        cell.font = bold_font

                # Make all cells in Suspicious sheet larger and bold for YES flags
                for row in ws.iter_rows(min_row=2):
                    for cell in row:
                        cell.font = Font(size=11)
                        if '⚠ YES' in str(cell.value or ''):
                            cell.font = red_bold_font

            # Style Merged Names sheet - highlight multiple IDs
            elif sheet_name == 'Merged Names':
                bold_font = Font(bold=True, size=12)
                orange_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

                # Bold header row
                for cell in ws[1]:
                    if cell.value:
                        cell.font = bold_font

                # Highlight rows with multiple IDs
                for row in ws.iter_rows(min_row=2):
                    for cell in row:
                        cell.font = Font(size=11)
                        # Check if this row has merged IDs (contains |)
                        if cell.column == 2 and '|' in str(cell.value or ''):
                            # Highlight the entire row
                            row_cells = ws[cell.row]
                            for c in row_cells:
                                c.fill = orange_fill
                                c.font = Font(bold=True, size=11)

    print(f'Exported to {filename}')


def main():
    """Main execution function"""
    print('Finding monthly files...')
    files = find_monthly_files()

    if not files:
        print('No monthly files found (XX.2568.xlsx)')
        return

    print(f'Found {len(files)} files: {", ".join(files)}')

    # Load and process all months
    all_months_data = []
    for filepath in files:
        try:
            print(f'  Processing {filepath}...')
            df = load_monthly_file(filepath)
            employees = extract_absence_data(df)
            all_months_data.append(employees)
            print(f'  -> {len(employees)} employees')
        except Exception as e:
            print(f'  Error: {e}')
            continue

    if not all_months_data:
        print('No data could be processed')
        return

    print(f'\nAggregating {len(all_months_data)} months...')
    aggregated = aggregate_yearly_totals(all_months_data)
    print(f'-> {len(aggregated)} employees')

    print(f'\nCalculating summary statistics...')
    summary_stats = calculate_summary_stats(aggregated, all_months_data)
    summary_df = pd.DataFrame(summary_stats)

    print(f'\nExporting...')
    df_output = create_output_dataframe(aggregated)
    export_to_excel(df_output, summary_df, aggregated, all_months_data)

    print(f'\nDone! {len(aggregated)} employees processed.')
    print(f'\nSummary sheet added to verify calculations:')


if __name__ == '__main__':
    main()
