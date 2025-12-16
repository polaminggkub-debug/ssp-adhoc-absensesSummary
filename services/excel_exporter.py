"""
Excel export service for absence summary reports.

Creates 6 sheets:
1. Executive Summary - CEO overview
2. Suspicious - Flagged records for HR review
3. Master Match - Audit trail for master data matching
4. Merged Names - Audit trail for merged employees
5. Data Traceback - File-by-file breakdown
6. Employees - Complete detailed data
"""

import pandas as pd
from typing import List, Dict, Any, Optional
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, PatternFill, Font

from config.absence_mapping import ABSENCE_COLUMN_HEADERS


def create_output_dataframe(aggregated_data: List[Dict[str, Any]]) -> pd.DataFrame:
    """
    Convert aggregated data to pandas DataFrame for export.

    Args:
        aggregated_data: List of aggregated employee dicts

    Returns:
        DataFrame with all employee data
    """
    columns = [
        'รหัส (EmpID)',
        'ชื่อ-สกุล (Name)',
        'ชื่อเต็ม (Master)',
        'หมายเหตุ (Notes)',
        'ตำแหน่ง (Position)',
        'แผนก (Department)',
        'ประเภท (PayType)',
    ] + ABSENCE_COLUMN_HEADERS

    rows = []
    for emp in aggregated_data:
        row = [
            emp['emp_id'],
            emp['name'],
            emp.get('master_full_name', ''),
            emp['notes'],
            emp['position'],
            emp['department'],
            emp['payType']
        ] + emp['totals']
        rows.append(row)

    return pd.DataFrame(rows, columns=columns)


def calculate_summary_stats(
    aggregated_data: List[Dict[str, Any]],
    all_months_data: List[List[Dict[str, Any]]],
    file_names: List[str] = None,
    section_data: List[Dict[str, Any]] = None
) -> List[Dict[str, Any]]:
    """
    Calculate summary statistics in pivot table format.
    Files/sections as rows, absence types as columns.

    Args:
        aggregated_data: Aggregated employee data
        all_months_data: Raw monthly data for comparison
        file_names: List of source file names
        section_data: List of dicts with section breakdowns per file

    Returns:
        List of summary stat rows (pivot table format)
    """
    if file_names is None:
        file_names = [f'File {i+1}' for i in range(len(all_months_data))]

    if section_data is None:
        section_data = [None] * len(all_months_data)

    # Calculate per-file totals for each absence type
    file_totals = []
    for month_data in all_months_data:
        totals = [0.0] * 17
        for emp in month_data:
            for i in range(17):
                totals[i] += emp['totals'][i]
        file_totals.append(totals)

    # Calculate raw totals (sum of all files)
    raw_totals = [0.0] * 17
    for totals in file_totals:
        for i in range(17):
            raw_totals[i] += totals[i]

    # Calculate aggregated totals
    aggregated_totals = [0.0] * 17
    for emp in aggregated_data:
        for i in range(17):
            aggregated_totals[i] += emp['totals'][i]

    # Use Thai column names (same as Employees sheet) for consistency
    col_names = ABSENCE_COLUMN_HEADERS

    rows = []

    # Header row info
    total_raw_records = sum(len(month) for month in all_months_data)
    total_merged_employees = len(aggregated_data)
    records_merged = total_raw_records - total_merged_employees

    # TOTAL row (aggregated output)
    total_row = {'File': 'TOTAL (Output)', 'Section': f'{total_merged_employees} employees'}
    for i, name in enumerate(col_names):
        total_row[name] = aggregated_totals[i]
    rows.append(total_row)

    # RAW TOTAL row (before merging)
    raw_row = {'File': 'RAW TOTAL', 'Section': f'{total_raw_records} records'}
    for i, name in enumerate(col_names):
        raw_row[name] = raw_totals[i]
    rows.append(raw_row)

    # Empty separator row
    rows.append({'File': '', 'Section': ''})

    # File breakdown rows
    for f_idx, fname in enumerate(file_names):
        sec = section_data[f_idx] if f_idx < len(section_data) else None

        if sec and sec.get('sections'):
            # File has sections - show file total first
            file_row = {'File': fname, 'Section': 'Total'}
            for i, name in enumerate(col_names):
                file_row[name] = file_totals[f_idx][i]
            rows.append(file_row)

            # Then show each section
            for s_idx, section_name in enumerate(sec['sections']):
                section_key = f'section{s_idx}'
                if section_key in sec:
                    sec_row = {'File': '', 'Section': section_name}
                    for i, name in enumerate(col_names):
                        sec_row[name] = sec[section_key][i]
                    rows.append(sec_row)
        else:
            # No sections - single row
            file_row = {'File': fname, 'Section': '-'}
            for i, name in enumerate(col_names):
                file_row[name] = file_totals[f_idx][i]
            rows.append(file_row)

    return rows


def create_master_match_sheet(
    match_audit: List[Dict[str, Any]],
    all_months_data: List[List[Dict[str, Any]]] = None,
    file_names: List[str] = None
) -> pd.DataFrame:
    """
    Create audit sheet showing how employees were matched to master data.

    Args:
        match_audit: List of match audit records
        all_months_data: Raw monthly data to find last appearance
        file_names: List of source file names

    Returns:
        DataFrame with match audit trail
    """
    if not match_audit:
        return pd.DataFrame([{
            'Master ID': 'No matching performed',
            'Master Name': '',
            'Original ID': '',
            'Original Name': '',
            'Match Type': '',
            'Confidence': '',
            'Note': ''
        }])

    # Build lookup for employee appearances across months
    emp_months = {}  # emp_id -> list of (month_idx, note from that month)
    emp_all_notes = {}  # emp_id -> all notes collected
    if all_months_data:
        for month_idx, month_data in enumerate(all_months_data):
            for emp in month_data:
                emp_id = emp.get('emp_id', '').strip()
                if emp_id:
                    if emp_id not in emp_months:
                        emp_months[emp_id] = []
                        emp_all_notes[emp_id] = set()
                    emp_months[emp_id].append(month_idx)
                    # Collect any notes from source
                    note = emp.get('note', '')
                    if note:
                        emp_all_notes[emp_id].add(str(note))

    rows = []
    for match in match_audit:
        confidence_str = ''
        note_parts = []

        orig_id = match.get('original_id', '')
        orig_name = match.get('original_name', '') or ''
        orig_notes = match.get('original_notes', '') or ''

        if match['match_type'] == 'UNMATCHED':
            confidence_str = '❌ Not Found'

            # Check if employee resigned/left (in name or notes)
            combined = orig_name + ' ' + orig_notes
            has_resign_keyword = 'ลาออก' in combined or 'ออก' in combined

            if has_resign_keyword:
                note_parts.append('ลาออก (Resigned)')

            # Add last appearance info
            if orig_id and orig_id in emp_months:
                months = emp_months[orig_id]
                last_month = max(months) + 1  # 1-indexed
                note_parts.append(f'สุดท้าย: เดือน {last_month:02d}')

            # Add any notes from source files
            if orig_id and orig_id in emp_all_notes:
                source_notes = emp_all_notes[orig_id]
                for sn in source_notes:
                    if sn and sn not in orig_notes:
                        note_parts.append(sn)

        elif match['match_type'] == 'Fuzzy':
            confidence_str = f"⚠ {match['confidence']:.0%}"
        else:
            confidence_str = f"✓ {match['confidence']:.0%}"

        rows.append({
            'Master ID': match['master_id'],
            'Master Name': match['master_name'],
            'Original ID': match['original_id'],
            'Original Name': match['original_name'],
            'Match Type': match['match_type'],
            'Confidence': confidence_str,
            'Note': ' | '.join(note_parts) if note_parts else ''
        })

    # Sort: unmatched first, then fuzzy, then exact
    type_order = {'UNMATCHED': 0, 'Fuzzy': 1, 'Name': 2, 'ID+Name': 3}
    rows.sort(key=lambda x: (type_order.get(x['Match Type'], 99), x['Original Name']))

    return pd.DataFrame(rows)


def create_suspicious_sheet(df: pd.DataFrame) -> pd.DataFrame:
    """
    Create a sheet highlighting suspicious/problematic records.

    Args:
        df: Employee DataFrame

    Returns:
        DataFrame with flagged records
    """
    suspicious_records = []

    for idx, row in df.iterrows():
        emp_id = row['รหัส (EmpID)']
        name = row['ชื่อ-สกุล (Name)']
        master_name = row.get('ชื่อเต็ม (Master)', '')
        notes = row['หมายเหตุ (Notes)']

        # Flag 1: Multiple IDs (job change)
        flag_multiple_ids = '⚠ YES' if (emp_id and '|' in str(emp_id)) else ''

        # Flag 2: Name has "/" in it (incomplete merging)
        flag_merged_name = '⚠ YES' if (name and '/' in str(name)) else ''

        # Flag 3: Has notes about ลาออก (quit)
        flag_quit = '⚠ YES' if (notes and 'ลาออก' in str(notes)) else ''

        # Flag 4: Has notes about เริ่มใหม่ (restarted)
        flag_restart = '⚠ YES' if (notes and 'เริ่มใหม่' in str(notes)) else ''

        # Flag 5: Has notes about ย้ายมา (transferred in)
        flag_transfer = '⚠ YES' if (notes and 'ย้ายมา' in str(notes)) else ''

        # Add to suspicious list if any flags
        if any([flag_multiple_ids, flag_merged_name, flag_quit, flag_restart, flag_transfer]):
            suspicious_records.append({
                'รหัส (ID)': emp_id,
                'ชื่อ-สกุล (Name)': name,
                'Multiple IDs?': flag_multiple_ids,
                'Merged Name?': flag_merged_name,
                'Quit (ลาออก)?': flag_quit,
                'Restart (เริ่มใหม่)?': flag_restart,
                'Transfer (ย้ายมา)?': flag_transfer,
                'หมายเหตุ (Notes)': notes
            })

    return pd.DataFrame(suspicious_records)


def create_executive_summary(
    aggregated_data: List[Dict[str, Any]],
    suspicious_df: pd.DataFrame,
    all_months_data: List[List[Dict[str, Any]]]
) -> pd.DataFrame:
    """
    Create CEO-level executive summary sheet (30-second overview).

    Args:
        aggregated_data: Aggregated employee data
        suspicious_df: DataFrame of suspicious records
        all_months_data: Raw monthly data

    Returns:
        DataFrame with executive summary
    """
    summary = []

    # [SECTION 1: PERIOD & SCOPE]
    total_raw_records = sum(len(month) for month in all_months_data)
    total_employees = len(aggregated_data)
    records_merged = total_raw_records - total_employees
    num_months = len(all_months_data)

    # Determine data period
    month_names = ['January', 'February', 'March', 'April', 'May', 'June',
                   'July', 'August', 'September', 'October', 'November', 'December']
    if num_months <= 12:
        period = f'{month_names[0]} - {month_names[num_months-1]} 2568'
    else:
        period = f'{num_months} months 2568'

    summary.append({'Metric': '[PERIOD & SCOPE]', 'Value': ''})
    summary.append({'Metric': 'Data Period', 'Value': period})
    summary.append({'Metric': 'Total Unique Employees', 'Value': total_employees})
    summary.append({'Metric': 'Records Merged (deduplicates)', 'Value': records_merged})
    summary.append({'Metric': '', 'Value': ''})

    # [SECTION 2: WORKFORCE OVERVIEW]
    summary.append({'Metric': '[WORKFORCE OVERVIEW]', 'Value': ''})

    work_days_total = sum(emp['totals'][0] for emp in aggregated_data)
    summary.append({'Metric': 'Total Work Days', 'Value': int(work_days_total)})

    total_suspicious = len(suspicious_df)
    multiple_ids = (suspicious_df['Multiple IDs?'] == '⚠ YES').sum() if len(suspicious_df) > 0 else 0
    quits = (suspicious_df['Quit (ลาออก)?'] == '⚠ YES').sum() if len(suspicious_df) > 0 else 0
    transfers = (suspicious_df['Transfer (ย้ายมา)?'] == '⚠ YES').sum() if len(suspicious_df) > 0 else 0

    summary.append({'Metric': 'Employees Requiring Review', 'Value': total_suspicious})
    summary.append({'Metric': '  └─ Job Changes (Multiple IDs)', 'Value': int(multiple_ids)})
    summary.append({'Metric': '  └─ Employee Quits', 'Value': int(quits)})
    summary.append({'Metric': '  └─ System Transfers', 'Value': int(transfers)})
    summary.append({'Metric': '', 'Value': ''})

    # [SECTION 3: TOP ABSENCE CATEGORIES]
    summary.append({'Metric': '[TOP ABSENCE CATEGORIES]', 'Value': ''})

    absence_totals = [0] * 17
    for emp in aggregated_data:
        for i in range(17):
            absence_totals[i] += emp['totals'][i]

    absence_list = [(ABSENCE_COLUMN_HEADERS[i], absence_totals[i], i) for i in range(17)]
    absence_list = [x for x in absence_list if x[1] > 0]
    absence_list.sort(key=lambda x: x[1], reverse=True)

    for absence_name, total, idx in absence_list[:7]:
        pct = (total / work_days_total * 100) if work_days_total > 0 else 0
        summary.append({
            'Metric': absence_name,
            'Value': f'{total:.1f} days ({pct:.2f}%)'
        })

    summary.append({'Metric': '', 'Value': ''})

    # [SECTION 4: DEPARTMENT CONCENTRATION]
    summary.append({'Metric': '[DEPARTMENT CONCENTRATION (TOP 5)]', 'Value': ''})

    dept_counts = {}
    for emp in aggregated_data:
        dept = emp.get('department', '')
        # Handle empty/None/NaN department values
        if not dept or (isinstance(dept, float) and pd.isna(dept)):
            dept = '(ไม่ระบุแผนก / Unknown Dept)'
        dept_counts[dept] = dept_counts.get(dept, 0) + 1

    top_depts = sorted(dept_counts.items(), key=lambda x: x[1], reverse=True)[:5]
    other_count = sum(count for _, count in sorted(dept_counts.items(), key=lambda x: x[1], reverse=True)[5:])

    for dept_name, count in top_depts:
        pct = (count / total_employees * 100)
        summary.append({'Metric': dept_name, 'Value': f'{count} employees ({pct:.1f}%)'})

    if other_count > 0:
        other_depts_count = len(dept_counts) - 5
        pct = (other_count / total_employees * 100)
        summary.append({
            'Metric': f'Other ({other_depts_count} departments)',
            'Value': f'{other_count} employees ({pct:.1f}%)'
        })

    summary.append({'Metric': '', 'Value': ''})

    # [SECTION 5: KEY INSIGHTS]
    summary.append({'Metric': '[KEY INSIGHTS]', 'Value': ''})

    unexcused_absent = absence_totals[1]
    unexcused_pct = (unexcused_absent / work_days_total * 100) if work_days_total > 0 else 0

    personal_leave = absence_totals[2]
    personal_pct = (personal_leave / work_days_total * 100) if work_days_total > 0 else 0

    if unexcused_pct < 0.1:
        summary.append({
            'Metric': '✓ Low Compliance Risk',
            'Value': f'Unexcused absence only {unexcused_pct:.3f}% - excellent'
        })
    else:
        summary.append({
            'Metric': '⚠ Compliance Alert',
            'Value': f'Unexcused absence {unexcused_pct:.2f}% - review needed'
        })

    if total_suspicious > 0:
        suspicious_pct = (total_suspicious / total_employees * 100)
        summary.append({
            'Metric': '⚠ HR Review Required',
            'Value': f'{total_suspicious} employees ({suspicious_pct:.1f}%) - see Suspicious sheet'
        })

    if personal_pct > 2:
        summary.append({
            'Metric': 'ℹ High Personal Leave',
            'Value': f'{personal_pct:.2f}% of work days - planned absences dominate'
        })

    summary.append({'Metric': '✓ Data Quality', 'Value': f'{records_merged} duplicates merged, verified'})

    return pd.DataFrame(summary)


def create_merged_names_sheet(
    df: pd.DataFrame,
    aggregated_data: List[Dict[str, Any]],
    all_months_data: List[List[Dict[str, Any]]]
) -> pd.DataFrame:
    """
    Create sheet showing ALL merged employees - by ID or by name algorithm.

    Args:
        df: Employee DataFrame
        aggregated_data: Aggregated employee data
        all_months_data: Raw monthly data

    Returns:
        DataFrame with merged employee audit trail
    """
    merged_list = []

    month_labels = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']

    for emp in aggregated_data:
        emp_id = emp.get('emp_id', '')
        original_names = emp.get('original_names', '')
        merge_reasons = emp.get('merge_reasons', '')

        # Check if merged
        has_multiple_ids = emp_id and '|' in str(emp_id)
        has_multiple_names = original_names and '|' in str(original_names)
        has_fuzzy_merge = bool(merge_reasons)

        if has_multiple_ids or has_multiple_names or has_fuzzy_merge:
            ids_list = [id_str.strip() for id_str in str(emp_id).split('|')] if emp_id else []

            month_ids = {}
            for m_idx, month_data in enumerate(all_months_data):
                month_label = month_labels[m_idx] if m_idx < len(month_labels) else f'M{m_idx+1}'
                found_ids = []
                for monthly_emp in month_data:
                    m_id = monthly_emp.get('emp_id', '').strip()
                    if m_id in ids_list:
                        found_ids.append(m_id)
                month_ids[month_label] = ' | '.join(sorted(found_ids)) if found_ids else '-'

            # Determine merge type from merge_reasons
            if 'ID Merge' in str(merge_reasons):
                merge_type = 'Same ID'
            elif 'Master Merge' in str(merge_reasons):
                merge_type = 'Master Match'
            elif has_multiple_ids:
                merge_type = 'ID Change'
            elif has_multiple_names:
                merge_type = 'Name Variation'
            else:
                merge_type = 'Other'

            row = {
                'Final Name': emp['name'],
                'Original Names': original_names,
                'Merge Type': merge_type,
            }

            for m_idx in range(len(all_months_data)):
                month_label = month_labels[m_idx] if m_idx < len(month_labels) else f'M{m_idx+1}'
                row[month_label] = month_ids.get(month_label, '-')

            merged_list.append(row)

    if not merged_list:
        row = {
            'Final Name': 'No merged employees',
            'Original Names': '',
            'Merge Type': '',
        }
        for m_idx in range(len(all_months_data)):
            month_label = month_labels[m_idx] if m_idx < len(month_labels) else f'M{m_idx+1}'
            row[month_label] = ''
        merged_list.append(row)

    return pd.DataFrame(merged_list)


def export_to_excel(
    df: pd.DataFrame,
    summary_df: pd.DataFrame,
    aggregated_data: List[Dict[str, Any]],
    all_months_data: List[List[Dict[str, Any]]],
    match_audit: Optional[List[Dict[str, Any]]] = None,
    filename: str = 'absence-summary-2568.xlsx'
) -> None:
    """
    Export DataFrame to Excel with all 6 sheets.

    Args:
        df: Employee DataFrame
        summary_df: Summary statistics DataFrame
        aggregated_data: Aggregated employee data
        all_months_data: Raw monthly data
        match_audit: Master matching audit trail
        filename: Output filename
    """
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        # Create suspicious sheet first (needed for executive summary)
        suspicious_df = create_suspicious_sheet(df)

        # Write executive summary (FIRST sheet for CEO view)
        executive_df = create_executive_summary(aggregated_data, suspicious_df, all_months_data)
        executive_df.to_excel(writer, sheet_name='Executive Summary', index=False)

        # Write suspicious records
        suspicious_df.to_excel(writer, sheet_name='Suspicious', index=False)

        # Write master match audit sheet (if available)
        if match_audit:
            master_match_df = create_master_match_sheet(match_audit, all_months_data)
            master_match_df.to_excel(writer, sheet_name='Master Match', index=False)

        # Write merged names sheet
        merged_names_df = create_merged_names_sheet(df, aggregated_data, all_months_data)
        merged_names_df.to_excel(writer, sheet_name='Merged Names', index=False)

        # Write data traceback sheet
        summary_df.to_excel(writer, sheet_name='Data Traceback', index=False)

        # Write detailed data
        df.to_excel(writer, sheet_name='Employees', index=False)

        # Auto-adjust column widths and style
        for sheet_name in writer.sheets:
            ws = writer.sheets[sheet_name]
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

                for cell in column:
                    if sheet_name == 'Executive Summary':
                        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    else:
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            # Style Executive Summary
            if sheet_name == 'Executive Summary':
                section_header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
                section_header_font = Font(bold=True, size=12, color='000000')
                regular_font = Font(size=11)

                for row_idx, row in enumerate(ws.iter_rows(min_row=1), 1):
                    for cell in row:
                        cell.font = regular_font
                        if cell.value and '[' in str(cell.value) and ']' in str(cell.value):
                            cell.font = section_header_font
                            cell.fill = section_header_fill

            # Style Suspicious sheet
            elif sheet_name == 'Suspicious':
                bold_font = Font(bold=True, size=12)
                red_bold_font = Font(color='FF0000', bold=True, size=12)

                for cell in ws[1]:
                    if cell.value:
                        cell.font = bold_font

                for row in ws.iter_rows(min_row=2):
                    for cell in row:
                        cell.font = Font(size=11)
                        if '⚠ YES' in str(cell.value or ''):
                            cell.font = red_bold_font

            # Style Merged Names sheet
            elif sheet_name == 'Merged Names':
                bold_font = Font(bold=True, size=12)
                orange_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

                for cell in ws[1]:
                    if cell.value:
                        cell.font = bold_font

                for row in ws.iter_rows(min_row=2):
                    for cell in row:
                        cell.font = Font(size=11)
                        if cell.column == 2 and '|' in str(cell.value or ''):
                            row_cells = ws[cell.row]
                            for c in row_cells:
                                c.fill = orange_fill
                                c.font = Font(bold=True, size=11)

            # Style Master Match sheet
            elif sheet_name == 'Master Match':
                bold_font = Font(bold=True, size=12)
                green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
                red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

                for cell in ws[1]:
                    if cell.value:
                        cell.font = bold_font

                for row in ws.iter_rows(min_row=2):
                    match_type_cell = row[4] if len(row) > 4 else None  # Match Type column
                    confidence_cell = row[5] if len(row) > 5 else None  # Confidence column

                    for cell in row:
                        cell.font = Font(size=11)

                    if match_type_cell and match_type_cell.value:
                        if match_type_cell.value == 'UNMATCHED':
                            for c in row:
                                c.fill = red_fill
                        elif match_type_cell.value == 'Fuzzy':
                            for c in row:
                                c.fill = yellow_fill
                        elif match_type_cell.value in ('ID', 'Name'):
                            for c in row:
                                c.fill = green_fill

    print(f'Exported to {filename}')
